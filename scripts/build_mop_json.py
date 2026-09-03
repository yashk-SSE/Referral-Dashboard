#!/usr/bin/env python3
"""
Build data/referral_mop.json from a monthly "MOP <Mon> Referral.xlsx" file.

WHY THIS EXISTS
---------------
referral_mop.json has always been maintained by hand (it is the one file in
data/ that no pipeline produces). From Sep'26 the source workbook carries a
3-way sub-channel split instead of the old 2-way BTL split, which is too many
numbers to retype safely every month -- hence this script.

SOURCE LAYOUT (Sep'26 onward)
-----------------------------
Row 1: metric group headers (BQL / MS / MD / Order / HOTO)
Row 2: sub-headers, 4 per metric: Total (Sales+Non-Sales) | Sales | Non Sales | BTL Referral
Row 3+: one row per city, col A = city name, "India" first.
A blank spacer column sits between metric groups, so each metric block starts
5 columns after the previous one (B, G, L, Q, V).

OUTPUT SCHEMA
-------------
Per city: sales / nonSales / btl / noBtl / combined, each {BQL,MS,MD,ORDER,HOTO}.

  noBtl    = the workbook's own "Total (Sales+Non-Sales)" column, taken AS GIVEN
             (not recomputed as sales+nonSales -- see ROUNDING below)
  combined = noBtl + btl

`noBtl`/`btl`/`combined` keep the exact names the pre-Sep'26 dashboard loader
already reads, so an old build of index.html consumes this file unchanged;
`sales`/`nonSales` are purely additive.

ROUNDING
--------
The workbook rounds every column independently, so "Total (Sales+Non-Sales)"
disagrees with sales+nonSales by +/-1 in ~19% of cells, and the India row
disagrees with the sum of its cities by 1-3. Both are inherent to the source
model, not transcription error. This script preserves the workbook's own
figures verbatim and reports the deltas rather than silently reconciling them
-- if that call ever changes, flip SUM_TOTALS below.
"""
import json, sys, argparse, datetime
from pathlib import Path
import openpyxl

METRICS = ['BQL', 'MS', 'MD', 'ORDER', 'HOTO']

# Dashboard variant key -> the JSON block that backs it. The dashboard's own
# MOP_VARIANTS uses 'all' for what this file calls 'combined'; everything else
# matches. Order here is the order the dashboard renders its split buttons in.
VARIANT_BLOCK = [('sales', 'sales'), ('nonSales', 'nonSales'), ('btl', 'btl'),
                 ('noBtl', 'noBtl'), ('all', 'combined')]

HISTORY_FILE = 'data/referral_mop_history.json'
# First column of each metric's 4-column block (1-indexed): B, G, L, Q, V
METRIC_COL = {'BQL': 2, 'MS': 7, 'MD': 12, 'ORDER': 17, 'HOTO': 22}
# Offsets within a metric block
OFF = {'noBtl': 0, 'sales': 1, 'nonSales': 2, 'btl': 3}

# False = trust the workbook's "Total (Sales+Non-Sales)" column (default).
# True  = recompute noBtl as sales+nonSales so the parts always add up.
SUM_TOTALS = False

CITY_RENAME = {'Bengaluru': 'Bangalore'}


def num(v):
    if v is None or v == '':
        return 0
    try:
        return int(round(float(v)))
    except (TypeError, ValueError):
        return 0


def parse(xlsx_path, sheet=None):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet] if sheet else wb.worksheets[0]

    # Sanity-check the header layout before trusting the hardcoded columns --
    # a shifted/renamed column would otherwise silently produce wrong targets.
    for m, col in METRIC_COL.items():
        got = str(ws.cell(row=1, column=col).value or '').strip().upper()
        if got != m.upper() and not (m == 'ORDER' and got == 'ORDER'):
            raise SystemExit(
                f'Unexpected layout: expected metric "{m}" in row 1 col {col}, found "{got}". '
                'The workbook column layout changed -- update METRIC_COL.')
        sub = str(ws.cell(row=2, column=col + OFF['btl']).value or '').strip().lower()
        if 'btl' not in sub:
            raise SystemExit(
                f'Unexpected layout: expected a BTL sub-header at row 2 col {col + OFF["btl"]}, '
                f'found "{sub}".')

    rows, warnings = [], []
    for r in range(3, ws.max_row + 1):
        raw = ws.cell(row=r, column=1).value
        if raw is None or not str(raw).strip():
            continue
        city = str(raw).strip()
        city = CITY_RENAME.get(city, city)

        blocks = {k: {} for k in ('sales', 'nonSales', 'btl', 'noBtl')}
        for m in METRICS:
            base = METRIC_COL[m]
            for key, off in OFF.items():
                blocks[key][m] = num(ws.cell(row=r, column=base + off).value)

        for m in METRICS:
            parts = blocks['sales'][m] + blocks['nonSales'][m]
            if SUM_TOTALS:
                blocks['noBtl'][m] = parts
            elif parts != blocks['noBtl'][m]:
                warnings.append((city, m, blocks['noBtl'][m], blocks['sales'][m],
                                 blocks['nonSales'][m], parts))

        blocks['combined'] = {m: blocks['noBtl'][m] + blocks['btl'][m] for m in METRICS}

        rows.append({'city': city, 'sales': blocks['sales'], 'nonSales': blocks['nonSales'],
                     'btl': blocks['btl'], 'noBtl': blocks['noBtl'],
                     'combined': blocks['combined']})
    return rows, warnings


def variants_present(rows):
    """Which dashboard variants this month actually has targets for.

    Carried in the history file per month so the Last Month Performance tab
    knows which split buttons to render -- Aug'26 has 3, Sep'26 has 5 -- instead
    of guessing or offering buttons that resolve to nothing.
    """
    out = []
    for vkey, block in VARIANT_BLOCK:
        if any(any((r.get(block) or {}).get(m, 0) for m in METRICS) for r in rows):
            out.append(vkey)
    return out


def month_label(ym):
    y, m = ym.split('-')
    return datetime.date(int(y), int(m), 1).strftime("%b'") + y[2:]


def write_history(rows, ym, path=HISTORY_FILE, quiet=False):
    """Add/replace one month in the month-keyed history file.

    referral_mop.json only ever holds the CURRENT month and is overwritten every
    time a new workbook lands, so last month's targets used to survive only in
    git. The Last Month Performance tab needs them at runtime, hence this file.
    It is written alongside -- never instead of -- the flat current-month file,
    so the live MOP tabs are unaffected.
    """
    p = Path(path)
    hist = {'months': {}}
    if p.exists():
        try:
            hist = json.loads(p.read_text(encoding='utf-8')) or {'months': {}}
        except ValueError:
            raise SystemExit(f'{path} exists but is not valid JSON -- refusing to overwrite it.')
    hist.setdefault('months', {})

    existed = ym in hist['months']
    if existed and not quiet:
        old = hist['months'][ym].get('cities', [])
        oldI = next((c for c in old if c.get('city') == 'India'), {})
        newI = next((c for c in rows if c.get('city') == 'India'), {})
        ob = (oldI.get('combined') or {}).get('BQL')
        nb = (newI.get('combined') or {}).get('BQL')
        print(f'\nNOTE: {ym} was already in {path} and is being replaced '
              f'(India combined BQL {ob} -> {nb}).')

    hist['months'][ym] = {'label': month_label(ym),
                          'variants': variants_present(rows),
                          'cities': rows}
    # Chronological, so the dashboard's month picker needs no sorting.
    hist['months'] = dict(sorted(hist['months'].items()))
    p.write_text(json.dumps(hist, indent=1) + '\n', encoding='utf-8')
    if not quiet:
        months = list(hist['months'])
        print(f'\n{"Replaced" if existed else "Added"} {ym} in {path} '
              f'({len(months)} months: {", ".join(months)})')
    return hist


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('xlsx')
    ap.add_argument('-o', '--out', default='data/referral_mop.json')
    ap.add_argument('--sheet', default=None)
    ap.add_argument('--dry-run', action='store_true')
    ap.add_argument('--month', default=None, metavar='YYYY-MM',
                    help='Month these targets apply to (default: the current '
                         'calendar month, since a MOP is loaded at the start of '
                         'the month it governs). Also the history file key.')
    ap.add_argument('--no-history', action='store_true',
                    help='Write only the flat current-month file, skipping '
                         + HISTORY_FILE + '.')
    a = ap.parse_args()

    ym = a.month or datetime.date.today().strftime('%Y-%m')
    try:
        datetime.datetime.strptime(ym, '%Y-%m')
    except ValueError:
        raise SystemExit(f'--month must look like YYYY-MM, got "{ym}".')

    rows, warnings = parse(a.xlsx, a.sheet)
    if not rows:
        raise SystemExit('No city rows parsed -- aborting rather than writing an empty MOP file.')
    if rows[0]['city'] != 'India':
        print(f'WARNING: first row is "{rows[0]["city"]}", not "India".', file=sys.stderr)

    print(f'Parsed {len(rows)} rows ({len(rows) - 1} cities + India) from {a.xlsx}')
    if warnings:
        print(f'\n{len(warnings)} cells where "Total (Sales+Non-Sales)" != sales+nonSales '
              f'(workbook rounding, preserved as given):')
        for city, m, tot, s, ns, parts in warnings:
            print(f'   {city:<12} {m:<6} Total={tot:<6} Sales={s:<6} NonSales={ns:<6} '
                  f'sum={parts:<6} ({parts - tot:+d})')

    india = rows[0]
    print('\nIndia targets:')
    print(f'  {"":<10} ' + ' '.join(f'{m:>7}' for m in METRICS))
    for k in ('sales', 'nonSales', 'btl', 'noBtl', 'combined'):
        print(f'  {k:<10} ' + ' '.join(f'{india[k][m]:>7}' for m in METRICS))

    print(f'\nMonth: {ym} ({month_label(ym)}) . variants present: '
          f'{", ".join(variants_present(rows))}')

    if a.dry_run:
        print('\n--dry-run: nothing written.')
        return
    Path(a.out).write_text(json.dumps(rows, indent=2) + '\n', encoding='utf-8')
    print(f'\nWrote {a.out} ({len(rows)} rows)')
    if not a.no_history:
        write_history(rows, ym)


if __name__ == '__main__':
    main()
